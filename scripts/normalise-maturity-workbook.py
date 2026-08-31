from pathlib import Path
import sys
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

source = Path(sys.argv[1]).resolve()
target = Path(sys.argv[2]).resolve()
target.parent.mkdir(parents=True, exist_ok=True)
source_book = load_workbook(source, data_only=True)

headers = [
    'Branch Code', 'Account No', 'Customer Name', 'Agent Name', 'Plan Amount',
    'Total Deposit Amount', 'Date of Joining', 'Maturity Date', 'Plan Name',
    'Current Maturity Amount', 'Tenure (Months)', 'Interest Rate', 'Source Sheet', 'Source Row'
]

def key(value):
    return ''.join(ch.lower() for ch in str(value or '') if ch.isalnum())

def pick(index, *names):
    for name in names:
        if key(name) in index:
            return index[key(name)]
    return None

def rows_from(sheet):
    header = {key(cell.value): i for i, cell in enumerate(sheet[1])}
    positions = {
        'account': pick(header, 'AccountNo', 'Account Number'),
        'customer': pick(header, 'Customer Name'),
        'agent': pick(header, 'Introducer(Agent Name)', 'Agent Name'),
        'plan': pick(header, 'Plan'),
        'deposit': pick(header, 'Total Deposit Amount'),
        'joined': pick(header, 'Date of Joining'),
        'maturity': pick(header, 'MaturityDate', 'Maturity Date'),
        'plan_name': pick(header, 'PlanName', 'Plan Name'),
        'maturity_amount': pick(header, 'MaturityAmount', 'Maturity Amount'),
        'current': pick(header, 'Current Maturity Amount'),
        'tenure': pick(header, 'Tenure (Months)'),
        'rate': pick(header, 'Interest Rate'),
    }
    output = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        customer = row[positions['customer']] if positions['customer'] is not None else None
        if not customer:
            continue
        def value(name, fallback=None):
            position = positions[name]
            return row[position] if position is not None and position < len(row) else fallback
        is_august = 'august' in sheet.title.lower()
        forecast_amount = value('maturity_amount', 0) if is_august else value('current', 0)
        source_maturity = value('maturity')
        forecast_maturity = datetime(2026, 8, 29) if is_august else source_maturity
        output.append([
            'AZM', str(value('account') or ''), str(customer).strip(), str(value('agent') or '').strip(),
            value('plan', 0) or 0, value('deposit', 0) or 0, value('joined'), forecast_maturity,
            str(value('plan_name') or '').strip(), forecast_amount or 0, value('tenure'),
            value('rate'), sheet.title, row_number,
        ])
    return output

groups = []
for sheet in source_book.worksheets:
    rows = rows_from(sheet)
    if rows:
        groups.append((sheet.title, rows))

book = Workbook()
summary = book.active
summary.title = 'Summary'
summary.sheet_view.showGridLines = False
summary['A1'] = 'MaturityFlow import pack'
summary['A1'].font = Font(name='Arial', size=18, bold=True, color='FFFFFF')
summary['A1'].fill = PatternFill('solid', fgColor='233876')
summary.merge_cells('A1:D1')
summary['A3'] = 'Branch Code'
summary['B3'] = 'AZM'
summary['A4'] = 'Source workbook'
summary['B4'] = source.name
summary['A6'] = 'Workbook sections'
summary['B6'] = 'Each month is kept on its own filtered worksheet.'

for index, (source_name, data_rows) in enumerate(groups, start=1):
    title = 'Current Month' if index == 1 else 'Next Month' if index == 2 else f'Month {index}'
    ws = book.create_sheet(title)
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='233876')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial', color='0000FF')
        row[4].number_format = '₹#,##0.00;[Red](₹#,##0.00);-'
        row[5].number_format = '₹#,##0.00;[Red](₹#,##0.00);-'
        row[9].number_format = '₹#,##0.00;[Red](₹#,##0.00);-'
        row[11].number_format = '0.00%'
        for position in (6, 7):
            if isinstance(row[position].value, datetime):
                row[position].number_format = 'dd-mm-yyyy'
    widths = [14, 20, 28, 26, 15, 20, 16, 16, 22, 23, 17, 15, 22, 12]
    for col, width in zip(ws.columns, widths):
        ws.column_dimensions[col[0].column_letter].width = width
    table = Table(displayName=f'Maturity{index}', ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    summary.cell(6 + index, 1, source_name)
    summary.cell(6 + index, 2, title)

summary['A11'] = 'Legend'
summary['A11'].font = Font(name='Arial', bold=True)
summary['A12'] = 'Blue text'
summary['B12'] = 'Source values copied from the supplied workbook; edit only if the source data is corrected.'
summary['A13'] = 'Branch assumption'
summary['B13'] = 'AZM — Azamgarh, specified by the user as the head branch.'
summary['A14'] = 'Purpose'
summary['B14'] = 'Upcoming-maturity forecast only. These rows are not payout cases until a form is received.'
summary['A15'] = 'August amount rule'
summary['B15'] = 'Uses MaturityAmount as the final amount: it was manually completed and already includes 8.50% interest.'
summary['A16'] = 'August maturity date'
summary['B16'] = 'All August maturity dates are set to 29-08-2026, as instructed by the user.'
for row in summary.iter_rows():
    for cell in row:
        if not cell.font.name:
            cell.font = Font(name='Arial')
summary.column_dimensions['A'].width = 24
summary.column_dimensions['B'].width = 75
summary.column_dimensions['C'].width = 26
summary.column_dimensions['D'].width = 4

book.calculation.fullCalcOnLoad = True
book.calculation.forceFullCalc = True
book.save(target)
print(target)
